import  openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]



for product_row in range(2,product_list.max_row + 1):
    supplier_name = product_list.cell(product_row,4).value
    #calculation number of products per  supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    #calculation total  value of inventory per supplier  should ask chatgpt below
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
         total_value_per_supplier[supplier_name] =   inventory * price

    #logic products with inverto

print(products_per_supplier)
print(total_value_per_supplier)
# print(products_per_supplier)
#
# {'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14}
# {'AAA Company': 10969059.95, 'BBB Company': 2375499.47, 'CCC Company': 8114363.62}
#


# "/home/evan/data/github/python/devops /.venv/spreadsheet/bin/python" /home/evan/data/github/python/spreadsheet/main.py
# adding a  new supplier
# adding a  new supplier
# adding a  new supplier
# {'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14}



    #
    # products_per_supplier["key"] = "value"
    # print(supplier_name.value)
    #

# pip install openpyxl
# 05 inventory supplier

#  pip  show openpyxl
# Name: openpyxl
# Version: 3.1.5
# Summary: A Python library to read/write Excel 2010 xlsx/xlsm files
# Home-page: https://openpyxl.readthedocs.io
# Author: See AUTHORS
# Author-email: charlie.clark@clark-consulting.eu
# License: MIT
# Location: /home/evan/data/github/python/devops /.venv/spreadsheet/lib/python3.11/site-packages

#sequence

# print(2,product_list.max_row)